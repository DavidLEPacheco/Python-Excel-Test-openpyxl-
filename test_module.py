'''
Created on Mar 10, 2019
This python file reads data from an excel file using openpyxl (must be installed on machine)
@author: Lionel PC
'''
import os
import openpyxl

os.chdir(r'''D:\Eclipse_Python_Javascript\Python Excel\test_package''') #Link to where your excel file is. Prefaced with 'r' which creates a raw string, otherwise a unicode error would arise. \u is a unicode escape function, and it is in this string (\Users). It is not expecting an s after the U.

wb=openpyxl.load_workbook('Test.xlsx')
print("Workbook opened")
print(wb.sheetnames) #Displays the sheet names of the workbook. This syntax replaces the old wb.get_sheet_names()

Names = wb['Names'] #Gets the sheet and stores it in a variable. This syntax replaces the old wb.get_sheet_by_name()
Dates = wb['Dates']

print(type(Names))
print(Names.title)
print(Names.max_row)
print(Names.max_column)
